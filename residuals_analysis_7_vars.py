import os
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import seaborn as sns
import scipy.stats as stats
import statsmodels.api as sm
from statsmodels.stats.diagnostic import het_breuschpagan
from statsmodels.stats.stattools import durbin_watson
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

INPUT_FILE  = "Data analytics group project.xlsx"
SHEET_NAME  = "Data-Sheet"
OUTPUT_DIR  = "residual_diagnostics"
DEPENDENT   = "P/B Ratio"
PREDICTORS  = ["ROE", "LN(Assets)"]
IGNORE      = {"Company", "Ticker", "SIC 3", "SIC 4"}
FIGSIZE     = (9, 6)
DPI         = 150

sns.set_theme(style="whitegrid", palette="muted")


def load_data():
    df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME)
    drop = [c for c in df.columns if c in IGNORE or c.startswith("Unnamed") or c.endswith(".1")]
    return df.drop(columns=drop, errors="ignore")


def prepare(df):
    sub = df[[DEPENDENT] + PREDICTORS].copy()
    for col in sub.columns:
        sub[col] = pd.to_numeric(sub[col], errors="coerce")
    return sub.dropna()


def fit(sub):
    y = sub[DEPENDENT]
    X = sm.add_constant(sub[PREDICTORS], has_constant="add")
    result = sm.OLS(y, X).fit()
    return result, y, X


def influence(result):
    inf = result.get_influence()
    return inf.cooks_distance[0], inf.resid_studentized_internal


def plot_resid_vs_fitted(fitted, resid):
    fig, ax = plt.subplots(figsize=FIGSIZE)
    ax.scatter(fitted, resid, alpha=0.6, edgecolors="steelblue", facecolors="lightsteelblue", s=50, linewidths=0.5)
    ax.axhline(0, color="red", linewidth=1.2, linestyle="--")
    lw = sm.nonparametric.lowess(resid, fitted, frac=0.6)
    ax.plot(lw[:, 0], lw[:, 1], color="darkorange", linewidth=2, label="LOWESS smoother")
    ax.set_xlabel("Fitted values", fontsize=12)
    ax.set_ylabel("Residuals", fontsize=12)
    ax.set_title("Residuals vs Fitted", fontsize=14, fontweight="bold")
    ax.legend()
    fig.tight_layout()
    fig.savefig(os.path.join(OUTPUT_DIR, "01_residuals_vs_fitted.png"), dpi=DPI)
    plt.close(fig)


def plot_histogram(resid):
    fig, ax = plt.subplots(figsize=FIGSIZE)
    sns.histplot(resid, kde=True, ax=ax, color="steelblue", edgecolor="white", bins=30, stat="density")
    mu, sigma = resid.mean(), resid.std()
    x = np.linspace(resid.min(), resid.max(), 300)
    ax.plot(x, stats.norm.pdf(x, mu, sigma), color="red", linewidth=2, linestyle="--", label="Normal fit")
    ax.set_xlabel("Residuals", fontsize=12)
    ax.set_ylabel("Density", fontsize=12)
    ax.set_title("Histogram of Residuals", fontsize=14, fontweight="bold")
    ax.legend()
    fig.tight_layout()
    fig.savefig(os.path.join(OUTPUT_DIR, "02_histogram_residuals.png"), dpi=DPI)
    plt.close(fig)


def plot_qq(resid):
    fig, ax = plt.subplots(figsize=FIGSIZE)
    (osm, osr), (slope, intercept, r) = stats.probplot(resid, dist="norm")
    ax.scatter(osm, osr, alpha=0.6, edgecolors="steelblue", facecolors="lightsteelblue", s=50, linewidths=0.5)
    x_line = np.array([min(osm), max(osm)])
    ax.plot(x_line, slope * x_line + intercept, color="red", linewidth=1.5, linestyle="--", label=f"Reference line (R={r:.4f})")
    ax.set_xlabel("Theoretical quantiles", fontsize=12)
    ax.set_ylabel("Sample quantiles", fontsize=12)
    ax.set_title("Normal Q-Q Plot of Residuals", fontsize=14, fontweight="bold")
    ax.legend()
    fig.tight_layout()
    fig.savefig(os.path.join(OUTPUT_DIR, "03_qq_plot.png"), dpi=DPI)
    plt.close(fig)


def plot_cooks(cooks_d):
    n = len(cooks_d)
    threshold = 4 / n
    indices = np.arange(1, n + 1)
    fig, ax = plt.subplots(figsize=FIGSIZE)
    ml, sl, _ = ax.stem(indices, cooks_d, linefmt="steelblue", markerfmt="o", basefmt=" ")
    ml.set_markersize(3)
    sl.set_linewidth(0.6)
    ax.axhline(threshold, color="red", linewidth=1.5, linestyle="--", label=f"Threshold 4/n = {threshold:.4f}")
    for idx in np.where(cooks_d > threshold)[0]:
        ax.annotate(str(idx + 1), (indices[idx], cooks_d[idx]), textcoords="offset points", xytext=(4, 4), fontsize=7)
    ax.set_xlabel("Observation index", fontsize=12)
    ax.set_ylabel("Cook's Distance", fontsize=12)
    ax.set_title("Cook's Distance", fontsize=14, fontweight="bold")
    ax.legend()
    fig.tight_layout()
    fig.savefig(os.path.join(OUTPUT_DIR, "04_cooks_distance.png"), dpi=DPI)
    plt.close(fig)


def plot_scale_location(fitted, std_resids):
    sqrt_abs = np.sqrt(np.abs(std_resids))
    fig, ax = plt.subplots(figsize=FIGSIZE)
    ax.scatter(fitted, sqrt_abs, alpha=0.6, edgecolors="steelblue", facecolors="lightsteelblue", s=50, linewidths=0.5)
    lw = sm.nonparametric.lowess(sqrt_abs, fitted, frac=0.6)
    ax.plot(lw[:, 0], lw[:, 1], color="darkorange", linewidth=2, label="LOWESS smoother")
    ax.set_xlabel("Fitted values", fontsize=12)
    ax.set_ylabel(r"$\sqrt{|\mathrm{Standardised\ residuals}|}$", fontsize=12)
    ax.set_title("Scale-Location Plot", fontsize=14, fontweight="bold")
    ax.legend()
    fig.tight_layout()
    fig.savefig(os.path.join(OUTPUT_DIR, "05_scale_location.png"), dpi=DPI)
    plt.close(fig)


def plot_std_residuals(std_resids):
    indices = np.arange(1, len(std_resids) + 1)
    fig, ax = plt.subplots(figsize=FIGSIZE)
    ax.scatter(indices, std_resids, alpha=0.6, edgecolors="steelblue", facecolors="lightsteelblue", s=50, linewidths=0.5)
    ax.axhline(0,  color="black",  linewidth=1)
    ax.axhline(2,  color="orange", linewidth=1.2, linestyle="--", label="±2σ")
    ax.axhline(-2, color="orange", linewidth=1.2, linestyle="--")
    ax.axhline(3,  color="red",    linewidth=1.2, linestyle=":",  label="±3σ")
    ax.axhline(-3, color="red",    linewidth=1.2, linestyle=":")
    for idx in np.where(np.abs(std_resids) > 3)[0]:
        ax.annotate(str(indices[idx]), (indices[idx], std_resids[idx]), textcoords="offset points", xytext=(4, 4), fontsize=7, color="red")
    ax.set_xlabel("Observation index", fontsize=12)
    ax.set_ylabel("Standardised residuals", fontsize=12)
    ax.set_title("Standardised Residuals", fontsize=14, fontweight="bold")
    ax.legend()
    fig.tight_layout()
    fig.savefig(os.path.join(OUTPUT_DIR, "06_standardised_residuals.png"), dpi=DPI)
    plt.close(fig)


def export_excel(sub, fitted, resid, std_resids, cooks_d, bp_lm, bp_pval, bp_f, bp_fpval, dw):
    data_df = pd.DataFrame({
        "obs_index":           sub.index,
        DEPENDENT:             sub[DEPENDENT].values,
        "ROE":                 sub["ROE"].values,
        "LN_Assets":           sub["LN(Assets)"].values,
        "fitted_values":       fitted,
        "residuals":           resid,
        "standardised_resids": std_resids,
        "cooks_distance":      cooks_d,
        "outlier_std_flag":    (np.abs(std_resids) > 2).astype(int),
        "influential_cook":    (cooks_d > 4 / len(sub)).astype(int),
    }).reset_index(drop=True)

    stats_df = pd.DataFrame([
        {"test": "Breusch-Pagan LM statistic", "value": bp_lm},
        {"test": "Breusch-Pagan p-value",       "value": bp_pval},
        {"test": "Breusch-Pagan F-statistic",   "value": bp_f},
        {"test": "Breusch-Pagan F p-value",     "value": bp_fpval},
        {"test": "Durbin-Watson statistic",      "value": dw},
    ])

    out_path = os.path.join(OUTPUT_DIR, "residuals_data.xlsx")
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        data_df.to_excel(writer,  sheet_name="Residuals Data",   index=False)
        stats_df.to_excel(writer, sheet_name="Test Statistics",  index=False)

    wb = load_workbook(out_path)
    for sheet_name, df in [("Residuals Data", data_df), ("Test Statistics", stats_df)]:
        ws = wb[sheet_name]
        for col_idx, col_name in enumerate(df.columns, start=1):
            if pd.api.types.is_float_dtype(df[col_name]):
                for cell in ws[get_column_letter(col_idx)][1:]:
                    cell.number_format = "0.000000000000"
    wb.save(out_path)


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    df  = load_data()
    sub = prepare(df)
    result, y, X = fit(sub)

    fitted     = result.fittedvalues.values
    resid      = result.resid.values
    cooks_d, std_resids = influence(result)

    bp_lm, bp_pval, bp_f, bp_fpval = het_breuschpagan(result.resid, X)
    dw = durbin_watson(result.resid)

    threshold = 4 / len(sub)
    print(result.summary())
    print(f"\nBreusch-Pagan  LM={bp_lm:.4f}  p={bp_pval:.4f}  F={bp_f:.4f}  p(F)={bp_fpval:.4f}")
    print(f"Durbin-Watson  DW={dw:.4f}")
    print(f"Outliers |std|>2 : {np.where(np.abs(std_resids)>2)[0]+1}")
    print(f"Influential Cook>4/n={threshold:.4f} : {np.where(cooks_d>threshold)[0]+1}")

    plot_resid_vs_fitted(fitted, resid)
    plot_histogram(resid)
    plot_qq(resid)
    plot_cooks(cooks_d)
    plot_scale_location(fitted, std_resids)
    plot_std_residuals(std_resids)

    export_excel(sub, fitted, resid, std_resids, cooks_d, bp_lm, bp_pval, bp_f, bp_fpval, dw)
    print(f"\nDone. Files saved to: {os.path.abspath(OUTPUT_DIR)}/")


if __name__ == "__main__":
    main()
