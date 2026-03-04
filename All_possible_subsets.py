"""
subset_regression.py
--------------------
Runs all 127 subset OLS regressions on the provided Excel dataset,
exports results to subset_regression_results.xlsx.

Usage:
    python subset_regression.py
"""

import itertools
import os
import pandas as pd
import numpy as np
import statsmodels.api as sm
from statsmodels.stats.outliers_influence import variance_inflation_factor


# ── Configuration ──────────────────────────────────────────────────────────────
INPUT_FILE  = "Data analytics group project.xlsx"
SHEET_NAME  = "Data-Sheet"
OUTPUT_FILE = "subset_regression_results.xlsx"

DEPENDENT   = "P/B Ratio"
PREDICTORS  = [
    "P/E Ratio",
    "LN(Assets)",
    "ROE",
    "SGrowth",
    "Debt/EBITDA Ratio",
    "D2834",
    "D2835",
]
IGNORE_COLS = [
    "Unnamed: 0", "Company", "Ticker", "SIC 3", "SIC 4",
    "P/E Ratio.1", "LN(Assets).1", "ROE.1", "SGrowth.1",
    "Debt/EBITDA Ratio.1", "D2834.1",
]
# ───────────────────────────────────────────────────────────────────────────────


def load_data(filepath: str, sheet: str) -> pd.DataFrame:
    """Load and clean the raw Excel sheet."""
    df = pd.read_excel(filepath, sheet_name=sheet, header=1)
    drop_cols = [c for c in IGNORE_COLS if c in df.columns]
    df = df.drop(columns=drop_cols)
    return df


def run_all_subsets(df: pd.DataFrame):
    """Run all 2^7 - 1 = 127 OLS regressions."""
    summary_rows   = []
    coef_rows      = []
    model_id       = 0

    for r in range(1, len(PREDICTORS) + 1):
        for subset in itertools.combinations(PREDICTORS, r):
            model_id += 1
            cols_needed = [DEPENDENT] + list(subset)

            sub_df = df[cols_needed].copy()
            for col in cols_needed:
                sub_df[col] = pd.to_numeric(sub_df[col], errors="coerce")
            sub_df = sub_df.dropna()

            y = sub_df[DEPENDENT]
            X = sm.add_constant(sub_df[list(subset)], has_constant="add")

            result = sm.OLS(y, X).fit()

            included_str = " + ".join(subset)

            summary_rows.append({
                "model_id":          model_id,
                "included_variables": included_str,
                "k":                 len(subset),
                "n_obs":             int(result.nobs),
                "R2":                result.rsquared,
                "Adjusted_R2":       result.rsquared_adj,
                "AIC":               result.aic,
                "BIC":               result.bic,
                "F_statistic":       result.fvalue,
                "Prob_F_stat":       result.f_pvalue,
            })

            for term in result.params.index:
                coef_rows.append({
                    "model_id":          model_id,
                    "included_variables": included_str,
                    "term":              term,
                    "coef":              result.params[term],
                    "std_err":           result.bse[term],
                    "t_value":           result.tvalues[term],
                    "p_value":           result.pvalues[term],
                })

    summary_df = pd.DataFrame(summary_rows)
    coef_df    = pd.DataFrame(coef_rows)
    return summary_df, coef_df


def compute_vif(df: pd.DataFrame) -> pd.DataFrame:
    """Compute VIF for the full model (all 7 predictors)."""
    sub_df = df[PREDICTORS + [DEPENDENT]].copy()
    for col in sub_df.columns:
        sub_df[col] = pd.to_numeric(sub_df[col], errors="coerce")
    sub_df = sub_df.dropna()

    X = sm.add_constant(sub_df[PREDICTORS], has_constant="add")
    vif_data = []
    for i, col in enumerate(X.columns):
        if col == "const":
            continue
        vif_val = variance_inflation_factor(X.values, i)
        vif_data.append({"variable": col, "VIF": vif_val})
    return pd.DataFrame(vif_data)


def build_top_models(summary_df: pd.DataFrame) -> pd.DataFrame:
    """Build TopModels sheet: top 20 by Adj R2, AIC, BIC."""
    top_adjr2 = summary_df.nlargest(20, "Adjusted_R2").copy()
    top_adjr2.insert(0, "rank_by", "Adjusted_R2")

    top_aic = summary_df.nsmallest(20, "AIC").copy()
    top_aic.insert(0, "rank_by", "AIC (lowest)")

    top_bic = summary_df.nsmallest(20, "BIC").copy()
    top_bic.insert(0, "rank_by", "BIC (lowest)")

    return pd.concat([top_adjr2, top_aic, top_bic], ignore_index=True)


def export_to_excel(summary_df, coef_df, top_df, vif_df, output_path: str):
    """Write all sheets to Excel without truncating precision."""
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary",      index=False)
        coef_df.to_excel(   writer, sheet_name="Coefficients", index=False)
        top_df.to_excel(    writer, sheet_name="TopModels",    index=False)
        vif_df.to_excel(    writer, sheet_name="VIF",          index=False)

        from openpyxl.utils import get_column_letter
        num_fmt = "0.000000000000"

        for sheet_name, df in [
            ("Summary",      summary_df),
            ("Coefficients", coef_df),
            ("TopModels",    top_df),
            ("VIF",          vif_df),
        ]:
            ws = writer.sheets[sheet_name]
            for col_idx, col_name in enumerate(df.columns, start=1):
                if pd.api.types.is_float_dtype(df[col_name]):
                    col_letter = get_column_letter(col_idx)
                    for cell in ws[col_letter][1:]:
                        cell.number_format = num_fmt


def main():
    print(f"Loading data from '{INPUT_FILE}', sheet '{SHEET_NAME}' ...")
    df = load_data(INPUT_FILE, SHEET_NAME)
    print(f"  Rows loaded: {len(df)}")

    print("Running all 127 subset regressions ...")
    summary_df, coef_df = run_all_subsets(df)
    print(f"  Models run: {len(summary_df)}")

    top_df = build_top_models(summary_df)

    print("Computing VIF for full model ...")
    vif_df = compute_vif(df)

    print(f"Exporting results to '{OUTPUT_FILE}' ...")
    export_to_excel(summary_df, coef_df, top_df, vif_df, OUTPUT_FILE)

    output_path = os.path.abspath(OUTPUT_FILE)
    print("\n─────────────────────────────────────────────────────")
    print(f"Models run      : {len(summary_df)}")
    print(f"Output file     : {output_path}")
    print("\nTop 5 models by Adjusted R²:")
    top5 = summary_df.nlargest(5, "Adjusted_R2")[
        ["model_id", "included_variables", "k", "n_obs", "Adjusted_R2", "AIC", "BIC"]
    ]
    print(top5.to_string(index=False))
    print("─────────────────────────────────────────────────────\n")


if __name__ == "__main__":
    main()
