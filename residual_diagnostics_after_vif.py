import os
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import scipy.stats as stats
from scipy.stats import t as t_dist, f as f_dist
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

INPUT_FILE = "Data analytics group project.xlsx"
SHEET_NAME = "After VIFs Data-Sheet"
OUTPUT_DIR = "residual_diagnostics_after_vif"
DEPENDENT  = "P/B Ratio"
PREDICTORS = ["ROE^2"]
FIGSIZE    = (9, 6)
DPI        = 150


def load_data():
    df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME, header=0)
    df.columns = df.iloc[0].tolist()
    df = df.iloc[1:].reset_index(drop=True)
    ignore = {"Company", "Ticker", "SIC 3", "SIC 4"}
    drop = [c for c in df.columns if str(c) in ignore or str(c).startswith("Unnamed")
            or str(c) == "nan" or str(c).endswith(".1")]
    return df.drop(columns=drop, errors="ignore")


def prepare(df):
    sub = df[[DEPENDENT] + PREDICTORS].copy()
    for col in sub.columns:
        sub[col] = pd.to_numeric(sub[col], errors="coerce")
    return sub.dropna().reset_index(drop=True)


def fit(sub):
    y    = sub[DEPENDENT].values.astype(float)
    X    = np.column_stack([np.ones(len(y)), sub[PREDICTORS].values.astype(float)])
    n, k = X.shape

    beta, _, _, _ = np.linalg.lstsq(X, y, rcond=None)
    fitted = X @ beta
    resid  = y - fitted

    ss_res = float(resid @ resid)
    ss_tot = float(((y - y.mean()) ** 2).sum())
    s2     = ss_res / (n - k)
    XtXinv = np.linalg.inv(X.T @ X)

    bse    = np.sqrt(np.diag(s2 * XtXinv))
    tvals  = beta / bse
    pvals  = 2 * t_dist.sf(np.abs(tvals), df=n - k)

    h          = np.diag(X @ XtXinv @ X.T)
    std_resids = resid / (np.sqrt(s2 * (1 - h)) + 1e-15)
    cooks      = (std_resids ** 2 / k) * (h / (1 - h + 1e-15))

    e2         = resid ** 2
    bp_beta, _, _, _ = np.linalg.lstsq(X, e2, rcond=None)
    r2_bp      = float(((X @ bp_beta - e2.mean()) ** 2).sum()) / float(((e2 - e2.mean()) ** 2).sum())
    bp_lm      = n * r2_bp
    bp_pval    = float(stats.chi2.sf(bp_lm, df=k - 1))
    bp_f       = (r2_bp / (k - 1)) / ((1 - r2_bp) / (n - k))
    bp_fpval   = float(f_dist.sf(bp_f, k - 1, n - k))
    dw         = float(np.sum(np.diff(resid) ** 2) / ss_res)

    r2     = 1 - ss_res / ss_tot
    r2_adj = 1 - (ss_res / (n - k)) / (ss_tot / (n - 1))
    ll     = -n / 2 * np.log(2 * np.pi * ss_res / n) - n / 2
    aic    = -2 * ll + 2 * k
    bic    = -2 * ll + k * np.log(n)
    fval   = ((ss_tot - ss_res) / (k - 1)) / (ss_res / (n - k))
    fpval  = float(f_dist.sf(fval, k - 1, n - k))

    print(f"n={n}  R2={r2:.6f}  AdjR2={r2_adj:.6f}  F={fval:.4f}(p={fpval:.2e})")
    print(f"AIC={aic:.4f}  BIC={bic:.4f}")
    for nm, b, se, t, p in zip(["const"] + PREDICTORS, beta, bse, tvals, pvals):
        print(f"  {nm}: coef={b:.6f}  se={se:.6f}  t={t:.4f}  p={p:.6f}")
    print(f"Breusch-Pagan: LM={bp_lm:.4f} p={bp_pval:.4e}  F={bp_f:.4f} p={bp_fpval:.4e}")
    print(f"Durbin-Watson: {dw:.4f}")
    print(f"Outliers |z|>2: {list(np.where(np.abs(std_resids)>2)[0]+1)}")
    print(f"Influential Cook>4/n: {list(np.where(cooks>4/n)[0]+1)}")

    return fitted, resid, std_resids, cooks, bp_lm, bp_pval, bp_f, bp_fpval, dw


def lowess(x, y, frac=0.6):
    n = len(x)
    r = int(np.ceil(frac * n))
    h = [np.sort(np.abs(x - x[i]))[r] for i in range(n)]
    w = (1 - np.clip(np.abs((x[:, None] - x[None, :]) / h), 0, 1) ** 3) ** 3
    yest = np.array([np.linalg.solve(
        [[np.sum(w[i]), np.sum(w[i]*x)], [np.sum(w[i]*x), np.sum(w[i]*x*x)]],
        [np.sum(w[i]*y), np.sum(w[i]*y*x)]
    ) @ [1, x[i]] for i in range(n)])
    order = np.argsort(x)
    return x[order], yest[order]


def plot_resid_vs_fitted(fitted, resid):
    fig, ax = plt.subplots(figsize=FIGSIZE)
    ax.scatter(fitted, resid, alpha=0.6, edgecolors="steelblue", facecolors="lightsteelblue", s=50, linewidths=0.5)
    ax.axhline(0, color="red", linewidth=1.2, linestyle="--")
    lx, ly = lowess(fitted, resid)
    ax.plot(lx, ly, color="darkorange", linewidth=2, label="LOWESS smoother")
    ax.set_xlabel("Fitted values", fontsize=12)
    ax.set_ylabel("Residuals", fontsize=12)
    ax.set_title("Residuals vs Fitted", fontsize=14, fontweight="bold")
    ax.legend(); fig.tight_layout()
    fig.savefig(os.path.join(OUTPUT_DIR, "01_residuals_vs_fitted.png"), dpi=DPI)
    plt.close(fig)


def plot_histogram(resid):
    fig, ax = plt.subplots(figsize=FIGSIZE)
    ax.hist(resid, bins=30, density=True, color="steelblue", edgecolor="white", alpha=0.8)
    x = np.linspace(resid.min(), resid.max(), 300)
    ax.plot(x, stats.norm.pdf(x, resid.mean(), resid.std()), color="red", linewidth=2, linestyle="--", label="Normal fit")
    ax.set_xlabel("Residuals", fontsize=12)
    ax.set_ylabel("Density", fontsize=12)
    ax.set_title("Histogram of Residuals", fontsize=14, fontweight="bold")
    ax.legend(); fig.tight_layout()
    fig.savefig(os.path.join(OUTPUT_DIR, "02_histogram_residuals.png"), dpi=DPI)
    plt.close(fig)


def plot_qq(resid):
    fig, ax = plt.subplots(figsize=FIGSIZE)
    (osm, osr), (slope, intercept, r) = stats.probplot(resid, dist="norm")
    ax.scatter(osm, osr, alpha=0.6, edgecolors="steelblue", facecolors="lightsteelblue", s=50, linewidths=0.5)
    x_line = np.array([min(osm), max(osm)])
    ax.plot(x_line, slope * x_line + intercept, color="red", linewidth=1.5, linestyle="--", label=f"Reference line (R={r:.4f})")
    ax.set_xlabel("Theoretical quantiles", fontsize=12)
    ax.set_ylabel("Sample quantiles", fontsize=12)
    ax.set_title("Normal Q-Q Plot of Residuals", fontsize=14, fontweight="bold")
    ax.legend(); fig.tight_layout()
    fig.savefig(os.path.join(OUTPUT_DIR, "03_qq_plot.png"), dpi=DPI)
    plt.close(fig)


def plot_cooks(cooks_d, n):
    threshold = 4 / n
    indices   = np.arange(1, n + 1)
    fig, ax   = plt.subplots(figsize=FIGSIZE)
    ml, sl, _ = ax.stem(indices, cooks_d, linefmt="steelblue", markerfmt="o", basefmt=" ")
    ml.set_markersize(3); sl.set_linewidth(0.6)
    ax.axhline(threshold, color="red", linewidth=1.5, linestyle="--", label=f"Threshold 4/n = {threshold:.4f}")
    for idx in np.where(cooks_d > threshold)[0]:
        ax.annotate(str(idx + 1), (indices[idx], cooks_d[idx]), textcoords="offset points", xytext=(4, 4), fontsize=7)
    ax.set_xlabel("Observation index", fontsize=12)
    ax.set_ylabel("Cook's Distance", fontsize=12)
    ax.set_title("Cook's Distance", fontsize=14, fontweight="bold")
    ax.legend(); fig.tight_layout()
    fig.savefig(os.path.join(OUTPUT_DIR, "04_cooks_distance.png"), dpi=DPI)
    plt.close(fig)


def plot_scale_location(fitted, std_resids):
    sqrt_abs = np.sqrt(np.abs(std_resids))
    fig, ax  = plt.subplots(figsize=FIGSIZE)
    ax.scatter(fitted, sqrt_abs, alpha=0.6, edgecolors="steelblue", facecolors="lightsteelblue", s=50, linewidths=0.5)
    lx, ly = lowess(fitted, sqrt_abs)
    ax.plot(lx, ly, color="darkorange", linewidth=2, label="LOWESS smoother")
    ax.set_xlabel("Fitted values", fontsize=12)
    ax.set_ylabel(r"$\sqrt{|\mathrm{Standardised\ residuals}|}$", fontsize=12)
    ax.set_title("Scale-Location Plot", fontsize=14, fontweight="bold")
    ax.legend(); fig.tight_layout()
    fig.savefig(os.path.join(OUTPUT_DIR, "05_scale_location.png"), dpi=DPI)
    plt.close(fig)


def plot_std_residuals(std_resids):
    indices = np.arange(1, len(std_resids) + 1)
    fig, ax = plt.subplots(figsize=FIGSIZE)
    ax.scatter(indices, std_resids, alpha=0.6, edgecolors="steelblue", facecolors="lightsteelblue", s=50, linewidths=0.5)
    ax.axhline(0,  color="black",  linewidth=1)
    ax.axhline(2,  color="orange", linewidth=1.2, linestyle="--", label="±2σ")
    ax.axhline(-2, color="orange", linewidth=1.2, linestyle="--")
    ax.axhline(3,  color="red",    linewidth=1.2, linestyle=":",  label="±3σ")
    ax.axhline(-3, color="red",    linewidth=1.2, linestyle=":")
    for idx in np.where(np.abs(std_resids) > 3)[0]:
        ax.annotate(str(indices[idx]), (indices[idx], std_resids[idx]), textcoords="offset points", xytext=(4, 4), fontsize=7, color="red")
    ax.set_xlabel("Observation index", fontsize=12)
    ax.set_ylabel("Standardised residuals", fontsize=12)
    ax.set_title("Standardised Residuals", fontsize=14, fontweight="bold")
    ax.legend(); fig.tight_layout()
    fig.savefig(os.path.join(OUTPUT_DIR, "06_standardised_residuals.png"), dpi=DPI)
    plt.close(fig)


def export_excel(sub, fitted, resid, std_resids, cooks_d, bp_lm, bp_pval, bp_f, bp_fpval, dw):
    n = len(sub)
    data_df = pd.DataFrame({
        "obs_index":           sub.index + 1,
        DEPENDENT:             sub[DEPENDENT].values,
        **{p: sub[p].values for p in PREDICTORS},
        "fitted_values":       fitted,
        "residuals":           resid,
        "standardised_resids": std_resids,
        "cooks_distance":      cooks_d,
        "outlier_std_flag":    (np.abs(std_resids) > 2).astype(int),
        "influential_cook":    (cooks_d > 4 / n).astype(int),
    })
    stats_df = pd.DataFrame([
        {"test": "Breusch-Pagan LM statistic", "value": bp_lm},
        {"test": "Breusch-Pagan p-value",       "value": bp_pval},
        {"test": "Breusch-Pagan F-statistic",   "value": bp_f},
        {"test": "Breusch-Pagan F p-value",     "value": bp_fpval},
        {"test": "Durbin-Watson statistic",      "value": dw},
    ])
    out_path = os.path.join(OUTPUT_DIR, "residuals_data_after_vif.xlsx")
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        data_df.to_excel(writer,  sheet_name="Residuals Data",  index=False)
        stats_df.to_excel(writer, sheet_name="Test Statistics", index=False)
    wb = load_workbook(out_path)
    for sname, df in [("Residuals Data", data_df), ("Test Statistics", stats_df)]:
        ws = wb[sname]
        for col_idx, col_name in enumerate(df.columns, start=1):
            if pd.api.types.is_float_dtype(df[col_name]):
                for cell in ws[get_column_letter(col_idx)][1:]:
                    cell.number_format = "0.000000000000"
    wb.save(out_path)


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    sub = prepare(load_data())
    fitted, resid, std_resids, cooks_d, bp_lm, bp_pval, bp_f, bp_fpval, dw = fit(sub)
    n = len(sub)
    plot_resid_vs_fitted(fitted, resid)
    plot_histogram(resid)
    plot_qq(resid)
    plot_cooks(cooks_d, n)
    plot_scale_location(fitted, std_resids)
    plot_std_residuals(std_resids)
    export_excel(sub, fitted, resid, std_resids, cooks_d, bp_lm, bp_pval, bp_f, bp_fpval, dw)


if __name__ == "__main__":
    main()
