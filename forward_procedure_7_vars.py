"""
forward_selection.py
--------------------
Implements Forward Selection OLS regression on the dataset
"Data analytics group project.xlsx" (sheet: Data-Sheet).

Exports results to forward_selection_results.xlsx with three sheets:
  - Forward Steps   : one row per round summarising the chosen variable
  - Tested Models   : every candidate model tested in every round
  - Final Model     : full statsmodels OLS summary of the winning model

Usage:
    python forward_selection.py
"""

import os
import pandas as pd
import numpy as np
import statsmodels.api as sm
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ── Configuration ──────────────────────────────────────────────────────────────
INPUT_FILE  = "Data analytics group project.xlsx"
SHEET_NAME  = "Data-Sheet"
OUTPUT_FILE = "forward_selection_results.xlsx"

DEPENDENT  = "P/B Ratio"
PREDICTORS = [
    "P/E Ratio",
    "LN(Assets)",
    "ROE",
    "SGrowth",
    "Debt/EBITDA Ratio",
    "D2834",
    "D2835",
]
# Columns to drop before any processing
IGNORE_SUFFIXES = (".1",)
IGNORE_EXACT    = {"Company", "Ticker", "SIC 3", "SIC 4"}
# ───────────────────────────────────────────────────────────────────────────────


def load_data(filepath: str, sheet: str) -> pd.DataFrame:
    """Load the Excel sheet and drop irrelevant / duplicated columns."""
    df = pd.read_excel(filepath, sheet_name=sheet, header=1)

    drop_cols = []
    for col in df.columns:
        if col in IGNORE_EXACT:
            drop_cols.append(col)
        elif col.startswith("Unnamed"):
            drop_cols.append(col)
        elif any(col.endswith(sfx) for sfx in IGNORE_SUFFIXES):
            drop_cols.append(col)

    df = df.drop(columns=drop_cols, errors="ignore")
    return df


def fit_ols(df: pd.DataFrame, y_col: str, x_cols: list) -> sm.regression.linear_model.RegressionResultsWrapper:
    """
    Fit OLS with intercept for given y and x column names.
    Converts to numeric and drops rows with any NaN in the required columns.
    """
    needed = [y_col] + x_cols
    sub = df[needed].copy()
    for col in needed:
        sub[col] = pd.to_numeric(sub[col], errors="coerce")
    sub = sub.dropna()

    y = sub[y_col]
    X = sm.add_constant(sub[x_cols], has_constant="add")
    return sm.OLS(y, X).fit()


def forward_selection(df: pd.DataFrame):
    """
    Perform forward selection by Adjusted R².

    Returns
    -------
    steps_rows   : list of dicts — one per completed round
    tested_rows  : list of dicts — every candidate model tested
    final_result : statsmodels RegressionResults — final winning model
    final_vars   : list of str  — predictors in the final model
    """
    remaining   = list(PREDICTORS)   # variables not yet selected
    selected    = []                  # variables already in the model
    best_adjr2  = -np.inf             # tracker for the current model's adj R²

    steps_rows  = []
    tested_rows = []
    round_num   = 0

    while remaining:
        round_num  += 1
        round_best  = {"adjr2": -np.inf, "var": None, "result": None}

        # ── Test adding each remaining variable to the current set ────────────
        for candidate in remaining:
            trial_vars = selected + [candidate]
            result     = fit_ols(df, DEPENDENT, trial_vars)

            tested_rows.append({
                "round":            round_num,
                "variables_tested": " + ".join(trial_vars),
                "added_candidate":  candidate,
                "n_obs":            int(result.nobs),
                "R2":               result.rsquared,
                "Adjusted_R2":      result.rsquared_adj,
                "AIC":              result.aic,
                "BIC":              result.bic,
            })

            if result.rsquared_adj > round_best["adjr2"]:
                round_best = {
                    "adjr2":  result.rsquared_adj,
                    "var":    candidate,
                    "result": result,
                    "vars":   trial_vars,
                }

        # ── Stopping rule: stop if best candidate doesn't improve adj R² ──────
        if round_best["adjr2"] <= best_adjr2:
            print(f"  Round {round_num}: Adding '{round_best['var']}' "
                  f"does NOT improve Adjusted R² "
                  f"({round_best['adjr2']:.8f} ≤ {best_adjr2:.8f}). Stopping.")
            break

        # ── Accept the best candidate ─────────────────────────────────────────
        best_adjr2 = round_best["adjr2"]
        chosen_var = round_best["var"]
        chosen_res = round_best["result"]
        selected.append(chosen_var)
        remaining.remove(chosen_var)

        print(f"  Round {round_num}: Added '{chosen_var}' → "
              f"Adj R² = {best_adjr2:.8f}  |  Model: {' + '.join(selected)}")

        steps_rows.append({
            "round":             round_num,
            "variables_in_model": " + ".join(selected),
            "added_variable":    chosen_var,
            "n_obs":             int(chosen_res.nobs),
            "R2":                chosen_res.rsquared,
            "Adjusted_R2":       chosen_res.rsquared_adj,
            "AIC":               chosen_res.aic,
            "BIC":               chosen_res.bic,
        })

    # ── Final model: refit with the accepted variables ────────────────────────
    if selected:
        final_result = fit_ols(df, DEPENDENT, selected)
    else:
        raise RuntimeError("No variables were selected — check your data.")

    return steps_rows, tested_rows, final_result, selected


def summary_to_df(result) -> pd.DataFrame:
    """
    Convert a statsmodels OLS summary into a tidy DataFrame
    suitable for writing to Excel.
    """
    rows = []

    # ── Top-level scalar statistics ───────────────────────────────────────────
    scalars = {
        "Dep. Variable":   result.model.endog_names,
        "R-squared":       result.rsquared,
        "Adj. R-squared":  result.rsquared_adj,
        "F-statistic":     result.fvalue,
        "Prob (F-stat)":   result.f_pvalue,
        "AIC":             result.aic,
        "BIC":             result.bic,
        "No. Observations": int(result.nobs),
        "Df Residuals":    int(result.df_resid),
        "Df Model":        int(result.df_model),
    }
    for k, v in scalars.items():
        rows.append({"section": "Model Info", "term": k, "value": v,
                     "std_err": "", "t": "", "p": "",
                     "[0.025": "", "0.975]": ""})

    # ── Blank separator ───────────────────────────────────────────────────────
    rows.append({k: "" for k in ["section","term","value","std_err","t","p","[0.025","0.975]"]})

    # ── Coefficient table ─────────────────────────────────────────────────────
    conf = result.conf_int()
    for term in result.params.index:
        rows.append({
            "section": "Coefficients",
            "term":    term,
            "value":   result.params[term],
            "std_err": result.bse[term],
            "t":       result.tvalues[term],
            "p":       result.pvalues[term],
            "[0.025":  conf.loc[term, 0],
            "0.975]":  conf.loc[term, 1],
        })

    return pd.DataFrame(rows)


def set_number_format(ws, df: pd.DataFrame, num_fmt: str = "0.000000000000"):
    """Apply a high-precision number format to all float columns in a sheet."""
    for col_idx, col_name in enumerate(df.columns, start=1):
        if pd.api.types.is_float_dtype(df[col_name]):
            col_letter = get_column_letter(col_idx)
            for cell in ws[col_letter][1:]:   # skip header row
                cell.number_format = num_fmt


def export_results(steps_rows, tested_rows, final_result, selected, output_path):
    """Write all three sheets to the output Excel workbook."""
    steps_df  = pd.DataFrame(steps_rows)
    tested_df = pd.DataFrame(tested_rows)
    final_df  = summary_to_df(final_result)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        steps_df.to_excel( writer, sheet_name="Forward Steps",  index=False)
        tested_df.to_excel(writer, sheet_name="Tested Models",  index=False)
        final_df.to_excel( writer, sheet_name="Final Model",    index=False)

    # ── Apply number formatting via openpyxl ──────────────────────────────────
    wb = load_workbook(output_path)
    set_number_format(wb["Forward Steps"],  steps_df)
    set_number_format(wb["Tested Models"],  tested_df)
    # Final Model has mixed types; format only numeric value/std_err/t/p columns
    ws_final = wb["Final Model"]
    float_cols = [c for c in final_df.columns
                  if final_df[c].apply(lambda x: isinstance(x, float)).any()]
    for col_idx, col_name in enumerate(final_df.columns, start=1):
        if col_name in float_cols:
            col_letter = get_column_letter(col_idx)
            for cell in ws_final[col_letter][1:]:
                if isinstance(cell.value, float):
                    cell.number_format = "0.000000000000"

    wb.save(output_path)


def main():
    # ── Load data ──────────────────────────────────────────────────────────────
    print(f"Loading '{INPUT_FILE}', sheet '{SHEET_NAME}' ...")
    df = load_data(INPUT_FILE, SHEET_NAME)
    print(f"  Rows loaded: {len(df)}\n")

    # ── Forward selection ──────────────────────────────────────────────────────
    print("Running Forward Selection ...")
    steps_rows, tested_rows, final_result, selected = forward_selection(df)

    # ── Export ─────────────────────────────────────────────────────────────────
    print(f"\nExporting results to '{OUTPUT_FILE}' ...")
    export_results(steps_rows, tested_rows, final_result, selected, OUTPUT_FILE)

    # ── Console summary ────────────────────────────────────────────────────────
    print("\n══════════════════════════════════════════════════════")
    print(f"Output file : {os.path.abspath(OUTPUT_FILE)}")
    print(f"Final model : {DEPENDENT} ~ {' + '.join(selected)}")
    print(f"  n_obs     : {int(final_result.nobs)}")
    print(f"  R²        : {final_result.rsquared}")
    print(f"  Adj R²    : {final_result.rsquared_adj}")
    print(f"  AIC       : {final_result.aic}")
    print(f"  BIC       : {final_result.bic}")
    print("══════════════════════════════════════════════════════\n")
    print(final_result.summary())


if __name__ == "__main__":
    main()
