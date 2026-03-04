import os
import pandas as pd
import numpy as np
from scipy.stats import t as t_dist, f as f_dist
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

INPUT_FILE  = "Data analytics group project.xlsx"
SHEET_NAME  = "After VIFs Data-Sheet"
OUTPUT_FILE = "forward_selection_results_after_vif.xlsx"
DEPENDENT   = "P/B Ratio"

PREDICTORS = [
    "P/E Ratio", "LN(Assets)", "SGrowth", "Debt/EBITDA Ratio",
    "D2834", "D2835", "ROE^2", "SGrowth^2",
]

XI_LABELS    = {var: f"X{i+1}" for i, var in enumerate(PREDICTORS)}
IGNORE_EXACT = {"Company", "Ticker", "SIC 3", "SIC 4"}


def load_data():
    df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME, header=0)
    df.columns = df.iloc[0].tolist()
    df = df.iloc[1:].reset_index(drop=True)
    drop = [c for c in df.columns if str(c) in IGNORE_EXACT or str(c).startswith("Unnamed")
            or str(c) == "nan" or str(c).endswith(".1")]
    return df.drop(columns=drop, errors="ignore")


class OLSResult:
    def __init__(self, y, X, x_names):
        n, k = X.shape
        self.nobs, self.df_model, self.df_resid = n, k - 1, n - k

        beta, _, _, _ = np.linalg.lstsq(X, y, rcond=None)
        self.params       = pd.Series(beta, index=x_names)
        self.fittedvalues = pd.Series(X @ beta)
        self.resid        = pd.Series(y - X @ beta)

        ss_res = float(self.resid @ self.resid)
        ss_tot = float(((y - y.mean()) ** 2).sum())
        s2     = ss_res / self.df_resid
        XtXinv = np.linalg.inv(X.T @ X)

        self.rsquared     = 1 - ss_res / ss_tot
        self.rsquared_adj = 1 - (ss_res / self.df_resid) / (ss_tot / (n - 1))
        self.bse          = pd.Series(np.sqrt(np.diag(s2 * XtXinv)), index=x_names)
        self.tvalues      = self.params / self.bse
        self.pvalues      = pd.Series(2 * t_dist.sf(np.abs(self.tvalues.values), df=self.df_resid), index=x_names)

        t_crit   = t_dist.ppf(0.975, df=self.df_resid)
        self._conf = pd.DataFrame({0: self.params - t_crit * self.bse, 1: self.params + t_crit * self.bse})

        ll        = -n / 2 * np.log(2 * np.pi * ss_res / n) - n / 2
        self.aic  = -2 * ll + 2 * k
        self.bic  = -2 * ll + k * np.log(n)
        self.fvalue   = ((ss_tot - ss_res) / self.df_model) / (ss_res / self.df_resid)
        self.f_pvalue = float(f_dist.sf(self.fvalue, self.df_model, self.df_resid))

    def conf_int(self):
        return self._conf


def fit_ols(df, y_col, x_cols):
    sub = df[[y_col] + x_cols].copy().apply(pd.to_numeric, errors="coerce").dropna()
    y   = sub[y_col].values.astype(float)
    X   = np.column_stack([np.ones(len(y)), sub[x_cols].values.astype(float)])
    return OLSResult(y, X, ["const"] + x_cols)


def forward_selection(df):
    remaining, selected, best_adjr2 = list(PREDICTORS), [], -np.inf
    steps_rows, tested_rows, round_num = [], [], 0

    while remaining:
        round_num += 1
        round_best = {"adjr2": -np.inf}

        for candidate in remaining:
            result = fit_ols(df, DEPENDENT, selected + [candidate])
            tested_rows.append({
                "round": round_num, "variables_tested": " + ".join(selected + [candidate]),
                "added_candidate": candidate, "n_obs": int(result.nobs),
                "R2": result.rsquared, "Adjusted_R2": result.rsquared_adj,
                "AIC": result.aic, "BIC": result.bic,
            })
            if result.rsquared_adj > round_best["adjr2"]:
                round_best = {"adjr2": result.rsquared_adj, "var": candidate, "result": result}

        if round_best["adjr2"] <= best_adjr2:
            break

        best_adjr2 = round_best["adjr2"]
        chosen_var, chosen_res = round_best["var"], round_best["result"]
        selected.append(chosen_var)
        remaining.remove(chosen_var)

        label = chosen_var if round_num == 1 else f"{chosen_var} ({XI_LABELS[chosen_var]})"
        steps_rows.append({
            "round": round_num,
            "variables_in_model": "+".join(f"{v} ({XI_LABELS[v]})" for v in selected),
            "added_variable": label, "n_obs": int(chosen_res.nobs),
            "R2": chosen_res.rsquared, "Adjusted_R2": chosen_res.rsquared_adj,
            "AIC": chosen_res.aic, "BIC": chosen_res.bic,
        })

    if not selected:
        raise RuntimeError("No variables selected.")

    return steps_rows, tested_rows, fit_ols(df, DEPENDENT, selected), selected


def summary_to_df(result):
    rows = []
    for k, v in {"Dep. Variable": DEPENDENT, "R-squared": result.rsquared,
                 "Adj. R-squared": result.rsquared_adj, "F-statistic": result.fvalue,
                 "Prob(F-stat)": result.f_pvalue, "AIC": result.aic, "BIC": result.bic,
                 "No. Observations": int(result.nobs), "Df Residuals": int(result.df_resid),
                 "Df Model": int(result.df_model)}.items():
        rows.append({"section": "Model Info", "term": k, "value": v,
                     "std_err": "", "t": "", "p": "", "[0.025": "", "0.975]": ""})
    rows.append({k: "" for k in ["section", "term", "value", "std_err", "t", "p", "[0.025", "0.975]"]})
    conf = result.conf_int()
    for term in result.params.index:
        rows.append({"section": "Coefficients", "term": term, "value": result.params[term],
                     "std_err": result.bse[term], "t": result.tvalues[term], "p": result.pvalues[term],
                     "[0.025": conf.loc[term, 0], "0.975]": conf.loc[term, 1]})
    return pd.DataFrame(rows)


def set_fmt(ws, df, fmt="0.000000000000"):
    for col_idx, col_name in enumerate(df.columns, start=1):
        if pd.api.types.is_float_dtype(df[col_name]):
            for cell in ws[get_column_letter(col_idx)][1:]:
                cell.number_format = fmt


def export_results(steps_rows, tested_rows, final_result, output_path):
    steps_df, tested_df, final_df = pd.DataFrame(steps_rows), pd.DataFrame(tested_rows), summary_to_df(final_result)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        steps_df.to_excel(writer,  sheet_name="Forward Steps", index=False)
        tested_df.to_excel(writer, sheet_name="Tested Models", index=False)
        final_df.to_excel(writer,  sheet_name="Final Model",   index=False)
    wb = load_workbook(output_path)
    set_fmt(wb["Forward Steps"], steps_df)
    set_fmt(wb["Tested Models"], tested_df)
    for col_idx, col_name in enumerate(final_df.columns, start=1):
        if final_df[col_name].apply(lambda x: isinstance(x, float)).any():
            for cell in wb["Final Model"][get_column_letter(col_idx)][1:]:
                if isinstance(cell.value, float):
                    cell.number_format = "0.000000000000"
    wb.save(output_path)


def main():
    df = load_data()
    steps_rows, tested_rows, final_result, selected = forward_selection(df)
    export_results(steps_rows, tested_rows, final_result, OUTPUT_FILE)
    print(f"Final model : {DEPENDENT} ~ {' + '.join(selected)}")
    print(f"R2={final_result.rsquared:.8f}  AdjR2={final_result.rsquared_adj:.8f}  AIC={final_result.aic:.4f}  BIC={final_result.bic:.4f}")


if __name__ == "__main__":
    main()
